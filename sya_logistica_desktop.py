import os
import subprocess
import sys
import tkinter as tk
from datetime import datetime
from tkinter import messagebox, ttk
import pandas as pd
import requests
import openpyxl
from openpyxl.utils import get_column_letter

# Configuración DPI para Windows
if sys.platform == "win32":
    try:
        import ctypes
        # For Windows 10 version 1607 and later: PROCESS_PER_MONITOR_DPI_AWARE
        ctypes.windll.shcore.SetProcessDpiAwareness(2)
    except (ImportError, AttributeError):
        try:
            # For Windows 8.1 and earlier:
            ctypes.windll.user32.SetProcessDPIAware()
        except (ImportError, AttributeError):
            pass

# Configuración para el manejo de imágenes
try:
    from PIL import Image, ImageTk
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

# Constantes
API_BASE_URL = "http://34.67.103.132:5000/api/logistica"
REQUERIMIENTOS_FILENAME = "sya_logistica_requerimientos.xlsx"
BDD_FILENAME = "logistica_materiales.csv"


# Clase para manejar utilidades de rutas y archivos
class FileUtils:
    @staticmethod
    def resource_path(relative_path):
        """Obtiene la ruta absoluta de un recurso, funciona tanto en desarrollo como en producción."""
        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)

    @staticmethod
    def obtener_ruta_aplicacion():
        """Obtiene la ruta base de la aplicación."""
        if getattr(sys, 'frozen', False):
            return os.path.dirname(sys.executable)
        else:
            return os.path.dirname(os.path.abspath(__file__))

    @staticmethod
    def crear_carpeta_descargas():
        """Crea la carpeta de descargas si no existe y devuelve su ruta."""
        ruta_app = FileUtils.obtener_ruta_aplicacion()
        ruta_descargas = os.path.join(ruta_app, "descargas")
        if not os.path.exists(ruta_descargas):
            try:
                os.makedirs(ruta_descargas)
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo crear la carpeta 'descargas':\n{e}")
        return ruta_descargas

    @staticmethod
    def abrir_archivo(ruta_archivo):
        """Abre un archivo con la aplicación predeterminada del sistema."""
        try:
            if sys.platform == 'win32':
                os.startfile(ruta_archivo)
            else:
                # Para macOS y Linux usamos subprocess
                if sys.platform == 'darwin':  # macOS
                    cmd = ['open', ruta_archivo]
                else:  # Linux
                    cmd = ['xdg-open', ruta_archivo]
                subprocess.call(cmd)
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Error al abrir el archivo: {e}")
            return False


# Clase para manejar operaciones con el servidor
class APIClient:
    @staticmethod
    def descargar_archivo(url, ruta_destino, status_callback=None):
        """Descarga un archivo desde una URL y lo guarda en la ruta especificada."""
        try:
            if status_callback:
                status_callback("Descargando archivo...")
                
            response = requests.get(url, stream=True, timeout=30)
            response.raise_for_status()

            with open(ruta_destino, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)
                    
            return True
        except requests.exceptions.RequestException as e:
            if status_callback:
                status_callback("Error al descargar el archivo")
            messagebox.showerror("Error de Conexión", f"No se pudo conectar al servidor:\n{e}")
            return False
        except Exception as e:
            if status_callback:
                status_callback("Error al descargar el archivo")
            messagebox.showerror("Error", f"Ocurrió un error al descargar el archivo:\n{e}")
            return False

    @staticmethod
    def subir_archivo(url, ruta_archivo, status_callback=None):
        """Sube un archivo al servidor."""
        try:
            if status_callback:
                status_callback("Subiendo archivo...")
                
            with open(ruta_archivo, 'rb') as f:
                files = {'file': (os.path.basename(ruta_archivo), f)}
                response = requests.post(url, files=files, timeout=60)
                response.raise_for_status()
                
            return True
        except requests.exceptions.RequestException as e:
            if status_callback:
                status_callback("Error al subir el archivo")
            messagebox.showerror("Error de Conexión", f"No se pudo conectar al servidor:\n{e}")
            return False
        except Exception as e:
            if status_callback:
                status_callback("Error al subir el archivo")
            messagebox.showerror("Error", f"Ocurrió un error al subir el archivo:\n{e}")
            return False


# Clase para manejar operaciones de Excel
class ExcelUtils:
    @staticmethod
    def ordenar_excel_por_fecha(ruta_archivo, status_callback=None):
        """Ordena un archivo Excel por la columna fecha de forma descendente."""
        try:
            if status_callback:
                status_callback("Ordenando datos por fecha...")
                
            # Leer el archivo Excel
            df = pd.read_excel(ruta_archivo)
            
            # Hacer una copia de la columna fecha original
            df['Fecha_original'] = df['Fecha']
            
            # Convertir la columna 'Fecha' a formato datetime para ordenamiento
            df['Fecha'] = pd.to_datetime(df['Fecha'], errors='coerce')
            
            # Ordenar por fecha descendente (más reciente primero)
            df_ordenado = df.sort_values(by='Fecha', ascending=False)
            
            # Convertir las fechas a formato dd/mm/yyyy
            df_ordenado['Fecha'] = df_ordenado['Fecha'].dt.strftime('%d/%m/%Y')
            
            # Si algunas fechas no se pudieron convertir, usar los valores originales
            mascara_nulos = df_ordenado['Fecha'].isna()
            if mascara_nulos.any():
                df_ordenado.loc[mascara_nulos, 'Fecha'] = df_ordenado.loc[mascara_nulos, 'Fecha_original']
            
            # Eliminar la columna auxiliar
            df_ordenado = df_ordenado.drop('Fecha_original', axis=1)
            
            # Guardar el archivo ordenado (sobreescribir)
            df_ordenado.to_excel(ruta_archivo, index=False)
            return df_ordenado
        except Exception as e:
            if status_callback:
                status_callback("Error al ordenar el archivo por fecha")
            print(f"Error al ordenar el archivo: {e}")
            return None

    @staticmethod
    def ajustar_columnas(ruta_archivo, status_callback=None):
        """Ajusta el ancho de las columnas del archivo Excel."""
        try:
            if status_callback:
                status_callback("Ajustando anchos de columnas...")
                
            # Leer el DataFrame para obtener dimensiones
            df = pd.read_excel(ruta_archivo)
            
            # Abrir el archivo Excel con openpyxl
            wb = openpyxl.load_workbook(ruta_archivo)
            hoja = wb.active
            
            # Autoajustar ancho de columnas basado en el contenido
            for col in range(1, len(df.columns) + 1):
                col_letra = get_column_letter(col)
                # Establecer un ancho mínimo para cada columna
                max_length = 10
                
                # Calcular el ancho basado en el título de la columna
                column_title = str(hoja.cell(row=1, column=col).value)
                if len(column_title) > max_length:
                    max_length = len(column_title)
                
                # Calcular el ancho basado en el contenido de la columna
                for celda in range(2, min(20, hoja.max_row + 1)):  # Limitamos a 20 filas para optimizar
                    valor_celda = str(hoja.cell(row=celda, column=col).value)
                    if len(valor_celda) > max_length:
                        max_length = len(valor_celda)
                
                # Ajustar el ancho (añadimos un margen de 2 caracteres)
                hoja.column_dimensions[col_letra].width = max_length + 2
            
            # Guardar el archivo con los ajustes
            wb.save(ruta_archivo)
            
            if status_callback:
                status_callback("Columnas ajustadas correctamente")
            return True
        except Exception as e:
            if status_callback:
                status_callback("Error al ajustar anchos de columnas")
            print(f"Error al ajustar anchos de columnas: {e}")
            return False


# Clase de aplicación principal
class SyaLogisticaApp:
    def __init__(self, root):
        self.root = root
        self.root.title("S&A - Sistema de Logística")
        self.root.resizable(True, True)
        self.root.configure(bg='#f0f0f0')
        
        # Variables para almacenar rutas de archivos
        self.ultimo_archivo = None
        self.ultimo_archivo_bdd = None
        
        # Configuración del icono
        try:
            self.root.iconbitmap(FileUtils.resource_path("images/smontyaragon.ico"))
        except Exception:
            pass
        
        # Configurar estilos
        self.configurar_estilos()
        
        # Crear carpeta de descargas
        FileUtils.crear_carpeta_descargas()
        
        # Inicializar la interfaz
        self.inicializar_interfaz()
        
        # Ajustar el tamaño mínimo de la ventana
        root.update_idletasks()
        root.minsize(int(root.winfo_reqwidth()*1.30), root.winfo_reqheight())
    
    def configurar_estilos(self):
        """Configura los estilos de la interfaz."""
        style = ttk.Style(self.root)
        style.theme_use('clam')

        style.configure("Title.TLabel", font=("Helvetica", 16, "bold"), foreground="#333", background='#f0f0f0')

        style.configure("Base.TButton", font=("Helvetica", 14), padding=(15, 12), relief="flat", background="#e0e0e0", foreground="#333")
        style.map("Base.TButton",
                background=[("active", "#f0f0f0")],
                relief=[("active", "raised")]
                )

        style.configure("Descargar.TButton", parent="Base.TButton", background="#cce0ff", foreground="#003366")
        style.map("Descargar.TButton",
                background=[("active", "#b3d1ff")],
                foreground=[("active", "#003366")]
                )

        style.configure("Abrir.TButton", parent="Base.TButton", background="#90EE90", foreground="#006400") # Verde claro
        style.map("Abrir.TButton",
                background=[("active", "#7FFFD4")], # Acuamarine
                foreground=[("active", "#006400")]
                )

        style.configure("Subir.TButton", parent="Base.TButton", background="#FFB6C1", foreground="#A52A2A") # Rosa claro
        style.map("Subir.TButton",
                background=[("active", "#FFA07A")], # Salmón claro
                foreground=[("active", "#A52A2A")]
                )
    
    def inicializar_interfaz(self):
        """Inicializa todos los componentes de la interfaz."""
        # Título
        titulo = ttk.Label(self.root, text="S&A - Sistema de Logística", style="Title.TLabel")
        titulo.pack(pady=20)
        
        # Cargar logo
        self.cargar_logo()
        
        # Frame principal para botones
        frame_botones_principal = ttk.Frame(self.root)
        frame_botones_principal.pack(pady=20, fill='x', expand=True)
        
        # Sección Requerimientos
        frame_requerimientos = ttk.LabelFrame(frame_botones_principal, text="Gestión de Requerimientos", padding=(10, 5))
        frame_requerimientos.pack(pady=10, padx=10, fill='x')

        btn_descargar_requerimientos = ttk.Button(
            frame_requerimientos, 
            text="Descargar Requerimientos", 
            command=self.descargar_requerimientos, 
            style="Descargar.TButton"
        )
        btn_descargar_requerimientos.pack(side='left', padx=5, expand=True, fill='x')

        btn_abrir_excel = ttk.Button(
            frame_requerimientos, 
            text="Abrir Excel Requerimientos", 
            command=self.abrir_excel, 
            style="Abrir.TButton"
        )
        btn_abrir_excel.pack(side='left', padx=5, expand=True, fill='x')
        
        # Sección Base de Datos Materiales
        frame_bdd = ttk.LabelFrame(frame_botones_principal, text="Base de Datos Materiales (CSV)", padding=(10, 5))
        frame_bdd.pack(pady=10, padx=10, fill='x')

        btn_descargar_bdd = ttk.Button(
            frame_bdd, 
            text="Descargar BB.DD.", 
            command=self.descargar_bdd, 
            style="Descargar.TButton"
        )
        btn_descargar_bdd.pack(side='left', padx=5, pady=5, expand=True, fill='x')

        btn_abrir_bdd = ttk.Button(
            frame_bdd, 
            text="Abrir BB.DD.", 
            command=self.abrir_bdd, 
            style="Abrir.TButton"
        )
        btn_abrir_bdd.pack(side='left', padx=5, pady=5, expand=True, fill='x')

        btn_subir_bdd = ttk.Button(
            frame_bdd, 
            text="Subir BB.DD.", 
            command=self.subir_bdd, 
            style="Subir.TButton"
        )
        btn_subir_bdd.pack(side='left', padx=5, pady=5, expand=True, fill='x')
        
        # Sección General
        frame_general = ttk.LabelFrame(frame_botones_principal, text="General", padding=(10, 5))
        frame_general.pack(pady=10, padx=10, fill='x')

        btn_abrir_carpeta = ttk.Button(
            frame_general, 
            text="Abrir Carpeta Descargas", 
            command=self.abrir_carpeta_descargas, 
            style="Abrir.TButton"
        )
        btn_abrir_carpeta.pack(padx=5, expand=True, fill='x')
        
        # Etiqueta de estado
        self.status_label = ttk.Label(self.root, text="Listo", font=("Helvetica", 10), background="#f0f0f0")
        self.status_label.pack(side=tk.BOTTOM, pady=10)
    
    def cargar_logo(self):
        """Carga e inserta el logo de la empresa."""
        if PIL_AVAILABLE:
            try:
                image = Image.open(FileUtils.resource_path("images/smontyaragon.png"))
                photo = ImageTk.PhotoImage(image)
                label_imagen = ttk.Label(self.root, image=photo)
                label_imagen.image = photo
                label_imagen.pack(pady=15)
            except Exception:
                # Si no se puede cargar la imagen, mostrar un texto alternativo
                label_texto = ttk.Label(self.root, text="S&A Logística", font=("Helvetica", 20, "bold"), foreground="#0066cc")
                label_texto.pack(pady=15)
        else:
            # Si PIL no está disponible, mostrar un texto alternativo
            label_texto = ttk.Label(self.root, text="S&A Logística", font=("Helvetica", 20, "bold"), foreground="#0066cc")
            label_texto.pack(pady=15)
    
    def actualizar_estado(self, texto):
        """Actualiza el texto del label de estado."""
        self.status_label.config(text=texto)
        self.root.update()
    
    def descargar_requerimientos(self):
        """Descarga el archivo Excel de requerimientos desde el servidor."""
        self.actualizar_estado("Descargando requerimientos...")
        
        # Preparar ruta de destino
        ruta_descargas = FileUtils.crear_carpeta_descargas()
        ruta_archivo = os.path.join(ruta_descargas, REQUERIMIENTOS_FILENAME)
        
        # Descargar archivo
        url = f"{API_BASE_URL}/descargar-requerimientos"
        descarga_exitosa = APIClient.descargar_archivo(url, ruta_archivo, self.actualizar_estado)
        
        if not descarga_exitosa:
            return None
        
        # Procesar archivo descargado
        df = ExcelUtils.ordenar_excel_por_fecha(ruta_archivo, self.actualizar_estado)
        if df is not None:
            ExcelUtils.ajustar_columnas(ruta_archivo, self.actualizar_estado)
        
        # Actualizar estado y mostrar mensaje
        self.actualizar_estado(f"Archivo descargado: {REQUERIMIENTOS_FILENAME}")
        messagebox.showinfo("Descarga Completada", 
                           f"El archivo de requerimientos ha sido descargado correctamente.")
        
        # Guardar la ruta del último archivo descargado
        self.ultimo_archivo = ruta_archivo
        return ruta_archivo
    
    def abrir_excel(self):
        """Abre el último archivo Excel descargado."""
        try:
            if hasattr(self, 'ultimo_archivo') and self.ultimo_archivo and os.path.exists(self.ultimo_archivo):
                # Abrir el archivo con la aplicación predeterminada
                if FileUtils.abrir_archivo(self.ultimo_archivo):
                    self.actualizar_estado(f"Archivo abierto: {os.path.basename(self.ultimo_archivo)}")
            else:
                # Buscar el archivo más reciente en la carpeta de descargas
                ruta_descargas = FileUtils.crear_carpeta_descargas()
                archivos_excel = [
                    f for f in os.listdir(ruta_descargas) 
                    if f.startswith("sya_logistica_requerimientos") and f.endswith(".xlsx")
                ]

                if archivos_excel:
                    # Ordenar por fecha de modificación (más reciente primero)
                    archivos_excel.sort(key=lambda x: os.path.getmtime(os.path.join(ruta_descargas, x)), reverse=True)
                    self.ultimo_archivo = os.path.join(ruta_descargas, archivos_excel[0])

                    # Abrir el archivo
                    if FileUtils.abrir_archivo(self.ultimo_archivo):
                        self.actualizar_estado(f"Archivo abierto: {os.path.basename(self.ultimo_archivo)}")
                else:
                    # Si no hay archivo descargado, mostrar mensaje
                    messagebox.showinfo(
                        "Información",
                        "No hay archivo para abrir. Por favor, descargue primero los requerimientos."
                    )
                    self.actualizar_estado("No hay archivo para abrir")
        except Exception as e:
            self.actualizar_estado("Error al abrir el archivo")
            messagebox.showerror("Error", f"Ocurrió un error al abrir el archivo:\n{e}")
    
    def abrir_carpeta_descargas(self):
        """Abre la carpeta de descargas."""
        ruta_descargas = FileUtils.crear_carpeta_descargas()
        try:
            if FileUtils.abrir_archivo(ruta_descargas):
                self.actualizar_estado("Carpeta de descargas abierta")
        except Exception as e:
            self.actualizar_estado("Error al abrir la carpeta")
            messagebox.showerror("Error", f"Ocurrió un error al abrir la carpeta de descargas:\n{e}")
    
    def descargar_bdd(self):
        """Descarga el archivo CSV de la base de datos de materiales desde el servidor."""
        self.actualizar_estado("Descargando BB.DD. Materiales...")
        
        # Preparar ruta de destino
        ruta_descargas = FileUtils.crear_carpeta_descargas()
        ruta_archivo_bdd = os.path.join(ruta_descargas, BDD_FILENAME)
        
        # Descargar archivo
        url = f"{API_BASE_URL}/descargar-bdd"
        descarga_exitosa = APIClient.descargar_archivo(url, ruta_archivo_bdd, self.actualizar_estado)
        
        if not descarga_exitosa:
            return None
        
        self.actualizar_estado(f"Archivo descargado: {BDD_FILENAME}")
        messagebox.showinfo("Descarga Completada",
                           f"El archivo de base de datos de materiales ha sido descargado correctamente.")
        
        self.ultimo_archivo_bdd = ruta_archivo_bdd
        return ruta_archivo_bdd
    
    def abrir_bdd(self):
        """Abre el último archivo CSV de base de datos descargado."""
        try:
            if self.ultimo_archivo_bdd and os.path.exists(self.ultimo_archivo_bdd):
                if FileUtils.abrir_archivo(self.ultimo_archivo_bdd):
                    self.actualizar_estado(f"Archivo abierto: {os.path.basename(self.ultimo_archivo_bdd)}")
            else:
                ruta_descargas = FileUtils.crear_carpeta_descargas()
                archivo_potencial = os.path.join(ruta_descargas, BDD_FILENAME)

                if os.path.exists(archivo_potencial):
                    self.ultimo_archivo_bdd = archivo_potencial
                    if FileUtils.abrir_archivo(self.ultimo_archivo_bdd):
                        self.actualizar_estado(f"Archivo BB.DD. abierto: {os.path.basename(self.ultimo_archivo_bdd)}")
                else:
                    messagebox.showinfo(
                        "Información",
                        "No hay archivo BB.DD. para abrir. Por favor, descargue primero la BB.DD."
                    )
                    self.actualizar_estado("No hay archivo BB.DD. para abrir")
        except Exception as e:
            self.actualizar_estado("Error al abrir archivo BB.DD.")
            messagebox.showerror("Error", f"Ocurrió un error al abrir el archivo BB.DD.:\n{e}")
    
    def subir_bdd(self):
        """Sube el archivo CSV de la base de datos de materiales al servidor."""
        if not self.ultimo_archivo_bdd or not os.path.exists(self.ultimo_archivo_bdd):
            # Intentar encontrar el archivo si no está en la variable global
            ruta_descargas = FileUtils.crear_carpeta_descargas()
            archivo_potencial = os.path.join(ruta_descargas, BDD_FILENAME)
            if os.path.exists(archivo_potencial):
                self.ultimo_archivo_bdd = archivo_potencial
            else:
                messagebox.showerror(
                    "Error", 
                    f"No se encuentra el archivo '{BDD_FILENAME}'.\nPor favor, descárguelo o asegúrese de que exista en la carpeta de descargas."
                )
                self.actualizar_estado("Archivo BB.DD. no encontrado para subir")
                return

        self.actualizar_estado("Subiendo BB.DD. Materiales...")
        
        # Subir archivo
        url = f"{API_BASE_URL}/subir-bdd"
        subida_exitosa = APIClient.subir_archivo(url, self.ultimo_archivo_bdd, self.actualizar_estado)
        
        if subida_exitosa:
            self.actualizar_estado("BB.DD. Materiales subida correctamente.")
            messagebox.showinfo(
                "Carga Completada",
                "El archivo de base de datos de materiales ha sido subido correctamente."
            )


# Punto de entrada principal
def main():
    """Inicializa y ejecuta la aplicación principal."""
    root = tk.Tk()
    global app
    app = SyaLogisticaApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
