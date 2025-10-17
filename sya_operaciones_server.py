# sya_operaciones_server.py
import os
import logging
from datetime import datetime
import pandas as pd
import openpyxl
from flask import Flask, request, jsonify, send_file
import zipfile
from flask_cors import CORS
import chardet

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# Usar rutas absolutas
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "registros_trabajo.xlsx")
REQUERIMIENTOS_EXCEL_FILE = os.path.join(BASE_DIR, "requerimientos_obra.xlsx")
EXCEL_FILE_AREA_TECNICA = os.path.join(BASE_DIR, "registros_trabajo_area_tecnica.xlsx")
MATERIALES_CSV_PATH = os.path.join(BASE_DIR, "operaciones_materiales.csv")
EQUIPOS_CSV_PATH = os.path.join(BASE_DIR, "operaciones_equipos.csv")
VEHICULOS_CSV_PATH = os.path.join(BASE_DIR, "operaciones_vehiculos.csv")
PERSONAL_CSV_PATH = os.path.join(BASE_DIR, "operaciones_personal.csv")

# Archivo Excel y directorio para registros de choferes
REGISTROS_CHOFERES_EXCEL = os.path.join(BASE_DIR, "registros_choferes.xlsx")
FOTOS_VEHICULOS_DIR = os.path.join(BASE_DIR, "fotos_vehiculos")

# Ruta al archivo CSV de conductores y vehiculos
CONDUCTORES_CSV_PATH = os.path.join(BASE_DIR, "aem_conductores.csv")
VEHICULOS_INFO_CSV_PATH = os.path.join(BASE_DIR, "aem_vehiculos.csv")

# Archivos para el sistema de logística
LOGISTICA_EXCEL_FILE = os.path.join(BASE_DIR, "sya_logistica_requerimientos.xlsx")
LOGISTICA_MATERIALES_CSV_PATH = os.path.join(BASE_DIR, "logistica_materiales.csv")
LOGISTICA_CLIENTES_CSV_PATH = os.path.join(BASE_DIR, "logistica_clientes.csv")

# Configuración de logging
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

def read_csv_with_encoding_detection(file_path):
    """
    Lee un archivo CSV detectando automáticamente la codificación.
    Intenta primero UTF-8, luego iso-8859-1, y finalmente usa detección automática.
    """
    encodings_to_try = ['utf-8', 'iso-8859-1', 'latin1', 'cp1252']

    for encoding in encodings_to_try:
        try:
            df = pd.read_csv(file_path, encoding=encoding)
            logging.info(f"Archivo {file_path} leído exitosamente con codificación {encoding}")
            return df
        except UnicodeDecodeError:
            continue
        except Exception as e:
            logging.warning(f"Error al leer {file_path} con codificación {encoding}: {str(e)}")
            continue

    # Si ninguna codificación funciona, intentar detección automática
    try:
        with open(file_path, 'rb') as file:
            raw_data = file.read()
            result = chardet.detect(raw_data)
            detected_encoding = result['encoding']

        if detected_encoding:
            df = pd.read_csv(file_path, encoding=detected_encoding)
            logging.info(f"Archivo {file_path} leído exitosamente con codificación detectada: {detected_encoding}")
            return df
    except Exception as e:
        logging.error(f"Error en detección automática de codificación para {file_path}: {str(e)}")

    # Como último recurso, intentar con errores='ignore'
    try:
        df = pd.read_csv(file_path, encoding='utf-8', errors='ignore')
        logging.warning(f"Archivo {file_path} leído con UTF-8 ignorando errores")
        return df
    except Exception as e:
        logging.error(f"Error final al leer {file_path}: {str(e)}")
        raise e

def inicializar_excel():
    """Inicializa los archivos Excel si no existen."""
    # Inicializar Excel de Reporte Diario
    if not os.path.exists(EXCEL_FILE):
        logging.info(f"Creando archivo Excel de reporte diario en: {EXCEL_FILE}")
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Reporte Principal"
        headers_reporte = [
            "Fecha", "Código Obra", "Nombre Ingeniero",
            "Nombre Supervisor", "Actividad Principal",
            "Supervisor Presente", "Avance Diario",
            "Incidentes", "Plan Siguiente Día", "Observaciones"
        ]
        ws1.append(headers_reporte)
        ws2 = wb.create_sheet(title="Materiales Usados")
        headers_materiales = ["Fecha", "Código Obra", "Nombre Ingeniero"]
        ws2.append(headers_materiales)
        ws3 = wb.create_sheet(title="Equipos Usados")
        headers_equipos = ["Fecha", "Código Obra", "Nombre Ingeniero"]
        ws3.append(headers_equipos)
        ws4 = wb.create_sheet(title="Vehículos Usados")
        headers_vehiculos = ["Fecha", "Código Obra", "Nombre Ingeniero"]
        ws4.append(headers_vehiculos)
        ws5 = wb.create_sheet(title="Personal de Campo")
        headers_personal = ["Fecha", "Código Obra", "Nombre Ingeniero"]
        ws5.append(headers_personal)
        wb.save(EXCEL_FILE)
    else:
        logging.info(f"El archivo Excel de reporte diario ya existe en: {EXCEL_FILE}")

    # Inicializar Excel de Requerimientos
    if not os.path.exists(REQUERIMIENTOS_EXCEL_FILE):
        logging.info(f"Creando archivo Excel de requerimientos en: {REQUERIMIENTOS_EXCEL_FILE}")
        wb_req = openpyxl.Workbook()
        ws_req = wb_req.active
        ws_req.title = "Requerimientos"
        headers_requerimientos_inicial = ["Fecha", "Código Obra", "Nombre Ingeniero"]
        ws_req.append(headers_requerimientos_inicial)
        wb_req.save(REQUERIMIENTOS_EXCEL_FILE)
    else:
        logging.info(f"El archivo Excel de requerimientos ya existe en: {REQUERIMIENTOS_EXCEL_FILE}")

    # Inicializar Excel de Registros de Choferes
    if not os.path.exists(REGISTROS_CHOFERES_EXCEL):
        logging.info(f"Creando archivo Excel de registros de choferes en: {REGISTROS_CHOFERES_EXCEL}")
        wb_choferes = openpyxl.Workbook()
        ws_choferes = wb_choferes.active
        ws_choferes.title = "Registros"
        headers_choferes = [
            "Fecha", "Nombre del Chofer", "Vehículo", "Placa", "Fecha de Salida",
            "Hora de Salida", "Ubicación Inicial", "Kilometraje Inicial",
            "Observaciones Salida", "Fecha de Llegada", "Hora de Retorno",
            "Ubicación Final", "Kilometraje Final", "Observaciones Llegada"
        ]
        ws_choferes.append(headers_choferes)
        wb_choferes.save(REGISTROS_CHOFERES_EXCEL)
    else:
        logging.info(f"El archivo Excel de registros de choferes ya existe en: {REGISTROS_CHOFERES_EXCEL}")

    # Inicializar Excel de Logística
    if not os.path.exists(LOGISTICA_EXCEL_FILE):
        logging.info(f"Creando archivo Excel de logística en: {LOGISTICA_EXCEL_FILE}")
        wb_logistica = openpyxl.Workbook()
        ws_logistica = wb_logistica.active
        ws_logistica.title = "Requerimientos"

        # Definir cabeceras
        cabeceras = [
            "Fecha", "Solicitante", "Orden de Trabajo", "Cliente",
            "Cantidad", "Unidad", "Producto", "Stock", "Timestamp"
        ]

        for col_num, header in enumerate(cabeceras, 1):
            ws_logistica.cell(row=1, column=col_num).value = header

        wb_logistica.save(LOGISTICA_EXCEL_FILE)
        logging.info(f"Archivo Excel de logística creado exitosamente")
    else:
        logging.info(f"El archivo Excel de logística ya existe en: {LOGISTICA_EXCEL_FILE}")

    # Inicializar Excel de Área Técnica
    if not os.path.exists(EXCEL_FILE_AREA_TECNICA):
        logging.info(f"Creando archivo Excel de área técnica en: {EXCEL_FILE_AREA_TECNICA}")
        wb_area_tecnica = openpyxl.Workbook()
        
        # Hoja 1: Reporte Principal
        ws1 = wb_area_tecnica.active
        ws1.title = "Reporte Principal"
        headers_reporte = [
            "Fecha", "Código Obra", "Nombre Ingeniero",
            "Nombre Supervisor", "Actividad Principal",
            "Supervisor Presente", "Avance Diario",
            "Incidentes", "Plan Siguiente Día", "Observaciones"
        ]
        ws1.append(headers_reporte)
        
        # Hoja 2: Materiales Usados
        ws2 = wb_area_tecnica.create_sheet(title="Materiales Usados")
        headers_materiales = ["O.T", "FECHA", "MATERIAL", "UNIDAD", "CANTIDAD", "PRECIO UNIT.", "COSTO"]
        ws2.append(headers_materiales)
        
        # Hoja 3: Equipos Usados
        ws3 = wb_area_tecnica.create_sheet(title="Equipos Usados")
        headers_equipos = ["O.T", "FECHA", "EQUIPO", "CODIGO", "H.T", "COSTO/HORA", "COSTO"]
        ws3.append(headers_equipos)
        
        # Hoja 4: Vehículos Usados
        ws4 = wb_area_tecnica.create_sheet(title="Vehículos Usados")
        headers_vehiculos = ["O.T", "FECHA", "VEHICULO", "PLACA", "H.T", "COSTO/HORA", "COSTO"]
        ws4.append(headers_vehiculos)
        
        # Hoja 5: Personal de Campo
        ws5 = wb_area_tecnica.create_sheet(title="Personal de Campo")
        headers_personal = ["O.T", "FECHA", "PERSONAL", "CATEGORIA", "COSTO/H.T", "COSTO/H.E", "H.T", "H.E", "COSTO"]
        ws5.append(headers_personal)
        
        wb_area_tecnica.save(EXCEL_FILE_AREA_TECNICA)
        logging.info(f"Archivo Excel de área técnica creado exitosamente")
    else:
        logging.info(f"El archivo Excel de área técnica ya existe en: {EXCEL_FILE_AREA_TECNICA}")

    # Inicializar CSV de Logística Materiales
    if not os.path.exists(LOGISTICA_MATERIALES_CSV_PATH):
        logging.info(f"Creando archivo CSV de materiales de logística en: {LOGISTICA_MATERIALES_CSV_PATH}")
        try:
            # Crear un DataFrame vacío con las cabeceras esperadas
            df_materiales_logistica = pd.DataFrame(columns=['material', 'unidad'])
            df_materiales_logistica.to_csv(LOGISTICA_MATERIALES_CSV_PATH, index=False)
            logging.info(f"Archivo CSV de materiales de logística creado exitosamente con cabeceras.")
        except Exception as e:
            logging.error(f"No se pudo crear el archivo CSV de materiales de logística: {e}")
    else:
        logging.info(f"El archivo CSV de materiales de logística ya existe en: {LOGISTICA_MATERIALES_CSV_PATH}")

    # Crear directorio de fotos si no existe
    if not os.path.exists(FOTOS_VEHICULOS_DIR):
        os.makedirs(FOTOS_VEHICULOS_DIR)
        logging.info(f"Directorio de fotos creado: {FOTOS_VEHICULOS_DIR}")

def actualizar_cabeceras_materiales(ws, num_materiales):
    """Actualiza las cabeceras de la hoja de materiales."""
    headers = list(ws.rows)[0]
    num_headers_actuales = len(headers)

    ultimo_material = 0
    for header in headers:
        header_value = header.value
        if header_value and header_value.startswith("Material"):
            try:
                numero = int(header_value.split(" ")[1])
                ultimo_material = max(ultimo_material, numero)
            except ValueError:
                pass

    if ultimo_material >= num_materiales:
        return

    for i in range(ultimo_material + 1, num_materiales + 1):
        ws.cell(row=1, column=num_headers_actuales + 1, value=f"Material {i}")
        ws.cell(row=1, column=num_headers_actuales + 2, value=f"Unidad {i}")
        ws.cell(row=1, column=num_headers_actuales + 3, value=f"Cantidad {i}")
        num_headers_actuales += 3

def actualizar_cabeceras_equipos(ws, num_equipos):
    """Actualiza las cabeceras de la hoja de equipos."""
    headers = list(ws.rows)[0]
    num_headers_actuales = len(headers)

    ultimo_equipo = 0
    for header in headers:
        header_value = header.value
        if header_value and header_value.startswith("Equipo"):
            try:
                numero = int(header_value.split(" ")[1])
                ultimo_equipo = max(ultimo_equipo, numero)
            except ValueError:
                pass

    if ultimo_equipo >= num_equipos:
        return

    for i in range(ultimo_equipo + 1, num_equipos + 1):
        ws.cell(row=1, column=num_headers_actuales + 1, value=f"Equipo {i}")
        ws.cell(row=1, column=num_headers_actuales + 2, value=f"Cantidad {i}")
        ws.cell(row=1, column=num_headers_actuales + 3, value=f"Ubicación {i}")
        num_headers_actuales += 3

def actualizar_cabeceras_vehiculos(ws, num_vehiculos):
    """Actualiza las cabeceras de la hoja de vehículos."""
    headers = list(ws.rows)[0]
    num_headers_actuales = len(headers)

    ultimo_vehiculo = 0
    for header in headers:
        header_value = header.value
        if header_value and header_value.startswith("Vehículo"):
            try:
                numero = int(header_value.split(" ")[1])
                ultimo_vehiculo = max(ultimo_vehiculo, numero)
            except ValueError:
                pass

    if ultimo_vehiculo >= num_vehiculos:
        return

    for i in range(ultimo_vehiculo + 1, num_vehiculos + 1):
        ws.cell(row=1, column=num_headers_actuales + 1, value=f"Vehículo {i}")
        ws.cell(row=1, column=num_headers_actuales + 2, value=f"Placa {i}")
        ws.cell(row=1, column=num_headers_actuales + 3, value=f"Combustible {i}")
        num_headers_actuales += 3

def actualizar_cabeceras_personal(ws, num_personal):
    """Actualiza las cabeceras de la hoja de personal."""
    headers = list(ws.rows)[0]
    num_headers_actuales = len(headers)

    ultimo_personal = 0
    for header in headers:
        header_value = header.value
        if header_value and header_value.startswith("Personal"):
            try:
                numero = int(header_value.split(" ")[1])
                ultimo_personal = max(ultimo_personal, numero)
            except ValueError:
                pass

    if ultimo_personal >= num_personal:
        return

    for i in range(ultimo_personal + 1, num_personal + 1):
        ws.cell(row=1, column=num_headers_actuales + 1, value=f"Personal {i}")
        ws.cell(row=1, column=num_headers_actuales + 2, value=f"Categoría {i}")
        ws.cell(row=1, column=num_headers_actuales + 3, value=f"Horas extras {i}")
        num_headers_actuales += 3

def actualizar_cabeceras_requerimientos(ws, num_items):
    """Actualiza las cabeceras de la hoja de requerimientos."""
    headers = list(ws.rows)[0]
    num_headers_actuales = len(headers)

    ultimo_item = 0
    for header in headers:
        header_value = header.value
        if header_value and header_value.startswith("Artículo"):
            try:
                numero = int(header_value.split(" ")[1])
                ultimo_item = max(ultimo_item, numero)
            except ValueError:
                pass

    if ultimo_item >= num_items:
        return

    for i in range(ultimo_item + 1, num_items + 1):
        ws.cell(row=1, column=num_headers_actuales + 1, value=f"Artículo {i}")
        ws.cell(row=1, column=num_headers_actuales + 2, value=f"Unidad {i}")
        ws.cell(row=1, column=num_headers_actuales + 3, value=f"Cantidad {i}")
        num_headers_actuales += 3


def procesar_datos(datos):
    """Procesa los datos del reporte diario."""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws_reporte = wb["Reporte Principal"]
        ws_materiales = wb["Materiales Usados"]
        ws_equipos = wb["Equipos Usados"]
        ws_vehiculos = wb["Vehículos Usados"]
        ws_personal = wb["Personal de Campo"]

        # Preparar fila de datos para "Reporte Principal"
        fila_reporte = [
            datetime.strptime(datos.get('fecha', ''), '%d/%m/%Y').date(),
            datos.get('codigo_obra', ''),
            datos.get('nombre_ingeniero', ''),
            datos.get('nombre_supervisor', ''),
            datos.get('actividad_principal', ''),
            'Sí' if datos.get('supervisor_presente', False) else 'No',
            datos.get('avance_diario', ''),
            datos.get('incidentes', ''),
            datos.get('siguiente_dia', ''),
            datos.get('observaciones', '')
        ]
        ws_reporte.append(fila_reporte)

        # Procesar materiales usados
        materiales = datos.get('materiales_usados', [])
        actualizar_cabeceras_materiales(ws_materiales, len(materiales))
        fila_materiales = [
            datetime.strptime(datos.get('fecha', ''), '%d/%m/%Y').date(),
            datos.get('codigo_obra', ''),
            datos.get('nombre_ingeniero', '')
        ]
        for material in materiales:
            fila_materiales.extend([material['nombre'], material['unidad'], material['cantidad']])
        ws_materiales.append(fila_materiales)

        # Procesar equipos usados
        equipos = datos.get('equipos_usados', [])
        actualizar_cabeceras_equipos(ws_equipos, len(equipos))
        fila_equipos = [
            datetime.strptime(datos.get('fecha', ''), '%d/%m/%Y').date(),
            datos.get('codigo_obra', ''),
            datos.get('nombre_ingeniero', '')
        ]
        for equipo in equipos:
            fila_equipos.extend([equipo['nombre'], equipo['cantidad'], equipo['ubicacion']])
        ws_equipos.append(fila_equipos)

        # Procesar vehículos usados
        vehiculos = datos.get('vehiculos_usados', [])
        actualizar_cabeceras_vehiculos(ws_vehiculos, len(vehiculos))
        fila_vehiculos = [
            datetime.strptime(datos.get('fecha', ''), '%d/%m/%Y').date(),
            datos.get('codigo_obra', ''),
            datos.get('nombre_ingeniero', '')
        ]
        for vehiculo in vehiculos:
            fila_vehiculos.extend([vehiculo['nombre'], vehiculo['placa'], vehiculo['combustible']])
        ws_vehiculos.append(fila_vehiculos)

        # Procesar personal de campo
        personal_campo = datos.get('personal_de_campo', [])
        actualizar_cabeceras_personal(ws_personal, len(personal_campo))
        fila_personal = [
            datetime.strptime(datos.get('fecha', ''), '%d/%m/%Y').date(),
            datos.get('codigo_obra', ''),
            datos.get('nombre_ingeniero', '')
        ]
        for personal in personal_campo:
            fila_personal.extend([personal['nombre_completo'], personal['categoria'], personal['horas_extras']])
        ws_personal.append(fila_personal)

        wb.save(EXCEL_FILE)
        logging.info(f"Datos recibidos de {datos.get('nombre_ingeniero', 'Unknown')} procesados exitosamente")

    except Exception as e:
        logging.error(f"Error al procesar datos: {str(e)}")


def procesar_datos_area_tecnica(datos):
    """Procesa los datos del reporte diario para el área técnica."""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE_AREA_TECNICA)
        ws_reporte = wb["Reporte Principal"]
        ws_materiales = wb["Materiales Usados"]
        ws_equipos = wb["Equipos Usados"]
        ws_vehiculos = wb["Vehículos Usados"]
        ws_personal = wb["Personal de Campo"]

        # Fecha para todas las hojas
        fecha = datetime.strptime(datos.get('fecha', ''), '%d/%m/%Y').date()
        codigo_obra = datos.get('codigo_obra', '')
        
        # Preparar fila de datos para "Reporte Principal" (igual que el original)
        fila_reporte = [
            fecha,
            codigo_obra,
            datos.get('nombre_ingeniero', ''),
            datos.get('nombre_supervisor', ''),
            datos.get('actividad_principal', ''),
            'Sí' if datos.get('supervisor_presente', False) else 'No',
            datos.get('avance_diario', ''),
            datos.get('incidentes', ''),
            datos.get('siguiente_dia', ''),
            datos.get('observaciones', '')
        ]
        ws_reporte.append(fila_reporte)

        # Procesar materiales usados (una fila por material)
        materiales = datos.get('materiales_usados', [])
        for material in materiales:
            precio_unit = float(material.get('precio', 0))
            cantidad = float(material.get('cantidad', 0))
            costo = round(precio_unit * cantidad, 2)
            
            fila_material = [
                codigo_obra,                      # O.T
                fecha,                           # FECHA
                material.get('nombre', ''),      # MATERIAL
                material.get('unidad', ''),      # UNIDAD
                cantidad,                        # CANTIDAD
                precio_unit,                     # PRECIO UNIT.
                costo                           # COSTO
            ]
            ws_materiales.append(fila_material)

        # Procesar equipos usados (una fila por equipo)
        equipos = datos.get('equipos_usados', [])
        for equipo in equipos:
            costo_hora = float(equipo.get('costo_hora', 0))
            horas_trabajadas = float(equipo.get('cantidad', 0))  # En el reporte diario, 'cantidad' son las horas trabajadas
            costo = round(costo_hora * horas_trabajadas, 2)
            
            fila_equipo = [
                codigo_obra,                      # O.T
                fecha,                           # FECHA
                equipo.get('nombre', ''),        # EQUIPO
                equipo.get('codigo', ''),        # CODIGO
                horas_trabajadas,                # H.T
                costo_hora,                      # COSTO/HORA
                costo                           # COSTO
            ]
            ws_equipos.append(fila_equipo)

        # Procesar vehículos usados (una fila por vehículo)
        vehiculos = datos.get('vehiculos_usados', [])
        for vehiculo in vehiculos:
            costo_hora = float(vehiculo.get('costo_hora', 0))
            horas_trabajadas = float(vehiculo.get('placa', 0))  # En el reporte diario, 'placa' son las horas trabajadas (según tu popup)
            costo = round(costo_hora * horas_trabajadas, 2)
            
            fila_vehiculo = [
                codigo_obra,                      # O.T
                fecha,                           # FECHA
                vehiculo.get('nombre', ''),       # VEHICULO
                vehiculo.get('placa_real', ''),   # PLACA (la placa real del vehículo)
                horas_trabajadas,                 # H.T
                costo_hora,                       # COSTO/HORA
                costo                            # COSTO
            ]
            ws_vehiculos.append(fila_vehiculo)

        # Procesar personal de campo (una fila por personal)
        personal_campo = datos.get('personal_de_campo', [])
        for personal in personal_campo:
            costo_ht = float(personal.get('costo_hora', 0))
            costo_he = float(personal.get('costo_hora_extra', 0))
            horas_extras_input = float(personal.get('horas_extras', 0))
            
            # Lógica para manejar horas extras negativas
            if horas_extras_input < 0:
                # Si las horas extras son negativas, se restan de las 8 horas trabajadas
                horas_trabajadas = 8.0 + horas_extras_input  # Sumar porque horas_extras_input es negativo
                horas_extras = 0.0
            else:
                # Si las horas extras son positivas o cero, funcionamiento normal
                horas_trabajadas = 8.0
                horas_extras = horas_extras_input
            
            # Calcular el costo total
            costo_total = round((costo_ht * horas_trabajadas) + (costo_he * horas_extras), 2)
            
            fila_personal = [
                codigo_obra,                           # O.T
                fecha,                                # FECHA
                personal.get('nombre_completo', ''), # PERSONAL
                personal.get('categoria', ''),        # CATEGORIA
                costo_ht,                            # COSTO/H.T
                costo_he,                            # COSTO/H.E
                horas_trabajadas,                    # H.T
                horas_extras,                        # H.E
                costo_total                          # COSTO
            ]
            ws_personal.append(fila_personal)

        wb.save(EXCEL_FILE_AREA_TECNICA)
        logging.info(f"Datos del área técnica recibidos de {datos.get('nombre_ingeniero', 'Unknown')} procesados exitosamente")

    except Exception as e:
        logging.error(f"Error al procesar datos del área técnica: {str(e)}")

def procesar_requerimientos(datos):
    """Procesa los datos de requerimientos."""
    logging.info("Datos de requerimientos recibidos:")
    logging.info(datos)
    try:
        wb_req = openpyxl.load_workbook(REQUERIMIENTOS_EXCEL_FILE)
        ws_requerimientos = wb_req["Requerimientos"]

        requerimientos = datos.get('requerimientos', [])
        actualizar_cabeceras_requerimientos(ws_requerimientos, len(requerimientos))

        fila_requerimientos = [
            datetime.strptime(datos.get('fecha', ''), '%d/%m/%Y').date(),
            datos.get('codigo_obra', ''),
            datos.get('nombre_ingeniero', '')
        ]
        for req in requerimientos:
            fila_requerimientos.extend([req['nombre'], req['unidad'], req['cantidad']])
        ws_requerimientos.append(fila_requerimientos)

        wb_req.save(REQUERIMIENTOS_EXCEL_FILE)
        logging.info(f"Requerimientos recibidos de {datos.get('nombre_ingeniero', 'Unknown')} procesados exitosamente")

    except Exception as e:
        logging.exception(f"Error al procesar requerimientos: {str(e)}")

def descargar_excel_flask():
    """Descarga el archivo Excel principal."""
    try:
        logging.info(f"Intentando enviar archivo: {EXCEL_FILE}")
        return send_file(EXCEL_FILE, as_attachment=True)
    except Exception as e:
        logging.error(f"Error al generar descarga de Excel: {str(e)}")
        return str(e), 500

def descargar_excel_area_tecnica_flask():
    """Descarga el archivo Excel del área técnica."""
    try:
        logging.info(f"Intentando enviar archivo de área técnica: {EXCEL_FILE_AREA_TECNICA}")
        return send_file(EXCEL_FILE_AREA_TECNICA, as_attachment=True, download_name='registros_trabajo_area_tecnica.xlsx')
    except Exception as e:
        logging.error(f"Error al generar descarga de Excel de área técnica: {str(e)}")
        return str(e), 500

def descargar_requerimientos_excel_flask():
    """Descarga el archivo Excel de requerimientos."""
    try:
        logging.info(f"Intentando enviar archivo de requerimientos: {REQUERIMIENTOS_EXCEL_FILE}")
        return send_file(REQUERIMIENTOS_EXCEL_FILE, as_attachment=True, download_name='requerimientos_obra.xlsx')
    except Exception as e:
        logging.error(f"Error al generar descarga de Excel de requerimientos: {str(e)}")
        return str(e), 500

def procesar_logistica_requerimientos(datos):
    """Procesa los datos de requerimientos de logística y los guarda en el Excel con ordenamiento por timestamp."""
    try:
        # Asegurar que el archivo Excel existe
        if not os.path.exists(LOGISTICA_EXCEL_FILE):
            logging.info(f"Creando archivo Excel de logística: {LOGISTICA_EXCEL_FILE}")
            wb_logistica = openpyxl.Workbook()
            ws_logistica = wb_logistica.active
            ws_logistica.title = "Requerimientos"

            # Definir cabeceras (archivo se creará manualmente con estas 9 columnas)
            cabeceras = [
                "Fecha", "Solicitante", "Orden de Trabajo", "Cliente",
                "Cantidad", "Unidad", "Producto", "Stock", "Timestamp"
            ]

            for col_num, header in enumerate(cabeceras, 1):
                ws_logistica.cell(row=1, column=col_num).value = header

            wb_logistica.save(LOGISTICA_EXCEL_FILE)
            logging.info("Archivo Excel de logística creado exitosamente")

        # Cargar el archivo Excel existente
        wb = openpyxl.load_workbook(LOGISTICA_EXCEL_FILE)
        ws = wb["Requerimientos"]

        # Datos comunes para todos los productos
        fecha_original = datos.get('fecha', '')
        solicitante = datos.get('solicitante', '')
        orden_trabajo = datos.get('orden_trabajo', '')
        cliente = datos.get('cliente', '')
        
        # Crear timestamp actual para ordenamiento
        timestamp_actual = datetime.now()
        timestamp_str = timestamp_actual.strftime("%Y-%m-%d %H:%M:%S")

        # Recopilar nuevas filas para insertar
        nuevas_filas = []
        productos = datos.get('productos', [])
        for producto in productos:
            nueva_fila = [
                fecha_original,  # Mantener formato original por ahora
                solicitante,
                orden_trabajo,
                cliente,
                producto.get('cantidad', 0.0),
                producto.get('unidad', ''),
                producto.get('producto', ''),
                '',  # Stock vacío
                timestamp_str  # Timestamp para ordenamiento
            ]
            nuevas_filas.append(nueva_fila)

        # Leer todas las filas existentes (excepto cabecera)
        filas_existentes = []
        for row_num in range(2, ws.max_row + 1):
            fila = []
            for col_num in range(1, 10):  # 9 columnas exactamente
                cell_value = ws.cell(row=row_num, column=col_num).value
                fila.append(cell_value if cell_value is not None else '')
            filas_existentes.append(fila)

        # Combinar filas existentes con nuevas filas
        todas_las_filas = filas_existentes + nuevas_filas

        # Ordenar por timestamp (columna 9) de manera descendente (más recientes primero)
        todas_las_filas.sort(key=lambda x: x[8], reverse=True)

        # Limpiar el contenido del archivo (mantener solo cabeceras)
        for row_num in range(ws.max_row, 1, -1):
            ws.delete_rows(row_num)

        # Escribir todas las filas ordenadas
        for fila in todas_las_filas:
            ws.append(fila)

        # La columna Timestamp ahora es visible y reemplaza a "Adquirido"

        # Guardar el archivo Excel
        wb.save(LOGISTICA_EXCEL_FILE)
        logging.info(f"Requerimientos de logística recibidos de {solicitante} procesados exitosamente con ordenamiento por timestamp")
        return True
    except Exception as e:
        logging.exception(f"Error al procesar requerimientos de logística: {str(e)}")
        return False

def descargar_logistica_excel_flask():
    """Descarga el archivo Excel de logística."""
    try:
        logging.info(f"Intentando enviar archivo de logística: {LOGISTICA_EXCEL_FILE}")
        return send_file(LOGISTICA_EXCEL_FILE, as_attachment=True, download_name='sya_logistica_requerimientos.xlsx')
    except Exception as e:
        logging.error(f"Error al generar descarga de Excel de logística: {str(e)}")
        return str(e), 500

def descargar_bdd_logistica_flask():
    """Descarga el archivo CSV de la base de datos de materiales de logística."""
    try:
        if not os.path.exists(LOGISTICA_MATERIALES_CSV_PATH):
            logging.error(f"Archivo BDD logística no encontrado: {LOGISTICA_MATERIALES_CSV_PATH}")
            return jsonify({"error": "Archivo BDD de logística no encontrado en el servidor."}), 404
        logging.info(f"Intentando enviar archivo BDD de logística: {LOGISTICA_MATERIALES_CSV_PATH}")
        return send_file(LOGISTICA_MATERIALES_CSV_PATH, as_attachment=True, download_name='logistica_materiales.csv')
    except Exception as e:
        logging.error(f"Error al generar descarga de CSV BDD de logística: {str(e)}")
        return jsonify({"error": f"Error interno del servidor: {str(e)}"}), 500


def agregar_nuevo_material_csv(nombre_material, unidad):
    """Agrega un nuevo material al archivo CSV."""
    try:
        df = read_csv_with_encoding_detection(MATERIALES_CSV_PATH)
        nuevo_material = pd.DataFrame([{'nombre_material': nombre_material, 'unidad': unidad}])
        df = pd.concat([df, nuevo_material], ignore_index=True)
        df.to_csv(MATERIALES_CSV_PATH, index=False, encoding='utf-8')
        logging.info(f"Nuevo material '{nombre_material}' agregado a {MATERIALES_CSV_PATH}")
        return True
    except Exception as e:
        logging.error(f"Error al agregar nuevo material a CSV: {str(e)}")
        return False

def agregar_nuevo_equipo_csv(nombre_equipo, ubicacion):
    """Agrega un nuevo equipo al archivo CSV."""
    try:
        df = read_csv_with_encoding_detection(EQUIPOS_CSV_PATH)
        nuevo_equipo = pd.DataFrame([{'nombre_equipo': nombre_equipo, 'ubicacion': ubicacion}])
        df = pd.concat([df, nuevo_equipo], ignore_index=True)
        df.to_csv(EQUIPOS_CSV_PATH, index=False, encoding='utf-8')
        logging.info(f"Nuevo equipo '{nombre_equipo}' agregado a {EQUIPOS_CSV_PATH}")
        return True
    except Exception as e:
        logging.error(f"Error al agregar nuevo equipo a CSV: {str(e)}")
        return False

def agregar_nuevo_vehiculo_csv(nombre_vehiculo, placa, combustible):
    """Agrega un nuevo vehículo al archivo CSV."""
    try:
        df = read_csv_with_encoding_detection(VEHICULOS_CSV_PATH)
        nuevo_vehiculo = pd.DataFrame([{'nombre_vehiculo': nombre_vehiculo, 'placa': placa, 'combustible': combustible}])
        df = pd.concat([df, nuevo_vehiculo], ignore_index=True)
        df.to_csv(VEHICULOS_CSV_PATH, index=False, encoding='utf-8')
        logging.info(f"Nuevo vehículo '{nombre_vehiculo}' agregado a {VEHICULOS_CSV_PATH}")
        return True
    except Exception as e:
        logging.error(f"Error al agregar nuevo vehículo a CSV: {str(e)}")
        return False

def agregar_nuevo_personal_csv(apellido_paterno, apellido_materno, nombres, categoria):
    """Agrega un nuevo personal al archivo CSV."""
    try:
        df = read_csv_with_encoding_detection(PERSONAL_CSV_PATH)
        nuevo_personal = pd.DataFrame([{
            'AP. PATERNO': apellido_paterno,
            'AP. MATERNO': apellido_materno,
            'NOMBRES': nombres,
            'CATEGORIA': categoria
        }])
        df = pd.concat([df, nuevo_personal], ignore_index=True)
        df.to_csv(PERSONAL_CSV_PATH, index=False, encoding='utf-8')
        logging.info(f"Nuevo personal '{nombres} {apellido_paterno}' agregado a {PERSONAL_CSV_PATH}")
        return True
    except Exception as e:
        logging.error(f"Error al agregar nuevo personal a CSV: {str(e)}")
        return False


# Inicializar Excel al inicio
inicializar_excel()

# Rutas de la API
@app.route('/api/materiales', methods=['GET'])
def get_materiales():
    """Obtiene la lista de materiales."""
    try:
        if not os.path.exists(MATERIALES_CSV_PATH):
            return jsonify({"error": f"No se encontró el archivo de materiales en {MATERIALES_CSV_PATH}"}), 404
        df = read_csv_with_encoding_detection(MATERIALES_CSV_PATH)
        materiales = df.to_dict(orient='records')
        logging.info("Materiales cargados exitosamente.")
        return jsonify(materiales)
    except Exception as e:
        logging.error(f"Error al obtener materiales: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/equipos', methods=['GET'])
def get_equipos():
    """Obtiene la lista de equipos."""
    try:
        if not os.path.exists(EQUIPOS_CSV_PATH):
            return jsonify({"error": f"No se encontró el archivo de equipos en {EQUIPOS_CSV_PATH}"}), 404
        df = read_csv_with_encoding_detection(EQUIPOS_CSV_PATH)
        equipos = df.to_dict(orient='records')
        logging.info("Equipos cargados exitosamente.")
        return jsonify(equipos)
    except Exception as e:
        logging.error(f"Error al obtener equipos: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/vehiculos', methods=['GET'])
def get_vehiculos():
    """Obtiene la lista de vehículos."""
    try:
        if not os.path.exists(VEHICULOS_CSV_PATH):
            return jsonify({"error": f"No se encontró el archivo de vehículos en {VEHICULOS_CSV_PATH}"}), 404
        df = read_csv_with_encoding_detection(VEHICULOS_CSV_PATH)
        vehiculos = df.to_dict(orient='records')
        logging.info("Vehículos cargados exitosamente.")
        return jsonify(vehiculos)
    except Exception as e:
        logging.error(f"Error al obtener vehículos: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/personal', methods=['GET'])
def get_personal():
    """Obtiene la lista de personal."""
    try:
        if not os.path.exists(PERSONAL_CSV_PATH):
            return jsonify({"error": f"No se encontró el archivo de personal en {PERSONAL_CSV_PATH}"}), 404
        df = read_csv_with_encoding_detection(PERSONAL_CSV_PATH)
        personal = df.to_dict(orient='records')
        logging.info("Personal de campo cargado exitosamente.")
        return jsonify(personal)
    except Exception as e:
        logging.error(f"Error al obtener personal: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/materiales/new', methods=['POST'])
def new_material():
    """Agrega un nuevo material."""
    data = request.json
    if not data or 'nombre_material' not in data or 'unidad' not in data:
        return jsonify({"error": "Nombre de material y unidad son requeridos"}), 400
    if agregar_nuevo_material_csv(data['nombre_material'], data['unidad']):
        return jsonify({"status": "success"}), 201
    else:
        return jsonify({"error": "Error al agregar nuevo material"}), 500

@app.route('/api/equipos/new', methods=['POST'])
def new_equipo():
    """Agrega un nuevo equipo."""
    data = request.json
    if not data or 'nombre_equipo' not in data or 'ubicacion' not in data:
        return jsonify({"error": "Nombre de equipo y ubicacion son requeridos"}), 400
    if agregar_nuevo_equipo_csv(data['nombre_equipo'], data['ubicacion']):
        return jsonify({"status": "success"}), 201
    else:
        return jsonify({"error": "Error al agregar nuevo equipo"}), 500

@app.route('/api/vehiculos/new', methods=['POST'])
def new_vehiculo():
    """Agrega un nuevo vehículo."""
    data = request.json
    if not data or 'nombre_vehiculo' not in data or 'placa' not in data or 'combustible' not in data:
        return jsonify({"error": "Nombre de vehículo, placa y combustible son requeridos"}), 400
    if agregar_nuevo_vehiculo_csv(data['nombre_vehiculo'], data['placa'], data['combustible']):
        return jsonify({"status": "success"}), 201
    else:
        return jsonify({"error": "Error al agregar nuevo vehículo"}), 500

@app.route('/api/personal/new', methods=['POST'])
def new_personal():
    """Agrega nuevo personal."""
    data = request.json
    if not data or 'nombre_completo' not in data or 'categoria' not in data:
        return jsonify({"error": "Nombre completo y categoría de personal son requeridos"}), 400

    nombre_completo = data['nombre_completo']
    partes_nombre = nombre_completo.split(',')
    if len(partes_nombre) != 2:
        return jsonify({"error": "Formato de nombre incorrecto. Debe ser 'apellido_paterno apellido_materno, nombres'"}), 400

    apellidos_str, nombres = partes_nombre
    apellidos_partes = apellidos_str.strip().split()
    if not apellidos_partes:
        return jsonify({"error": "Apellido paterno requerido"}), 400
    apellido_paterno = apellidos_partes[0]
    apellido_materno = apellidos_partes[1] if len(apellidos_partes) > 1 else ''

    if agregar_nuevo_personal_csv(apellido_paterno, apellido_materno, nombres.strip(), data['categoria']):
        return jsonify({"status": "success"}), 201
    else:
        return jsonify({"error": "Error al agregar nuevo personal"}), 500


@app.route('/recibir-datos', methods=['POST'])
def recibir_datos():
    """Recibe los datos del reporte diario."""
    datos = request.json
    procesar_datos(datos)
    procesar_datos_area_tecnica(datos)  # También procesar para área técnica
    return jsonify({"status": "success"})

@app.route('/recibir-requerimientos', methods=['POST'])
def recibir_requerimientos_route():
    """Recibe los datos de requerimientos."""
    datos = request.json
    print("Datos recibidos en /recibir-requerimientos:", datos)
    procesar_requerimientos(datos)
    return jsonify({"status": "success"})

@app.route('/descargar-excel', methods=['GET'])
def descargar_excel_route():
    """Descarga el archivo Excel principal."""
    return descargar_excel_flask()

@app.route('/descargar-excel-area-tecnica', methods=['GET'])
def descargar_excel_area_tecnica_route():
    """Descarga el archivo Excel del área técnica."""
    return descargar_excel_area_tecnica_flask()

@app.route('/descargar-requerimientos-excel', methods=['GET'])
def descargar_requerimientos_excel_route():
    """Descarga el archivo Excel de requerimientos."""
    return descargar_requerimientos_excel_flask()

# Funciones y rutas para la app de choferes
def procesar_datos_choferes(data, files):
    """Procesa los datos del formulario de choferes o solo guarda fotos si se proporciona un row_idx."""
    try:
        wb_choferes = openpyxl.load_workbook(REGISTROS_CHOFERES_EXCEL)
        ws_choferes = wb_choferes.active

        nombre_chofer = data.get("nombre_chofer")
        placa = data.get("placa")
        tipo_formulario = data.get("tipo_formulario")
        row_idx = data.get("row_idx")  # Identificador de fila (opcional)

        # Función para generar el nombre de la subcarpeta
        def generar_nombre_subcarpeta(fecha_salida, nombre_chofer, placa):
            fecha_salida_filename = fecha_salida.replace("-", "")
            nombre_chofer_filename = nombre_chofer.lower().replace(" ", "-")
            placa_filename = placa.replace(" ", "-")
            return f"{fecha_salida_filename}_{nombre_chofer_filename}_{placa_filename}"

        # Si solo se envía row_idx y fotos (segunda solicitud para llegada)
        if row_idx and not tipo_formulario:
            try:
                row_idx = int(row_idx)
                if row_idx <= 1 or row_idx > ws_choferes.max_row:
                    return False, "Índice de fila inválido."

                # Obtener la fecha de salida desde el Excel (columna 5: Fecha de Salida)
                fecha_salida_excel = ws_choferes.cell(row=row_idx, column=5).value
                if isinstance(fecha_salida_excel, datetime):
                    fecha_salida_str = fecha_salida_excel.strftime("%Y-%m-%d")
                else:
                    fecha_salida_str = str(fecha_salida_excel)

                # Generar el nombre de la subcarpeta
                subcarpeta_nombre = generar_nombre_subcarpeta(fecha_salida_str, nombre_chofer, placa)
                subcarpeta_path = os.path.join(FOTOS_VEHICULOS_DIR, subcarpeta_nombre)

                # Crear la subcarpeta si no existe
                if not os.path.exists(subcarpeta_path):
                    os.makedirs(subcarpeta_path)
                    logging.info(f"Subcarpeta creada: {subcarpeta_path}")

                # Guardar las fotos de llegada en la subcarpeta
                for i in range(1, 5):
                    foto_key = f"foto_km_final_{i}"
                    foto_fin = files.get(foto_key)
                    if foto_fin:
                        original_extension = os.path.splitext(foto_fin.filename)[1] if foto_fin.filename else ".jpg"
                        filename_fin = f"{subcarpeta_nombre}_llegada_{i}{original_extension}"
                        path_fin = os.path.join(subcarpeta_path, filename_fin)
                        foto_fin.save(path_fin)
                        logging.info(f"Foto de fin {i} guardada en {path_fin} para fila {row_idx}")
                return True, "Fotos de llegada guardadas correctamente."
            except ValueError:
                return False, "Índice de fila debe ser un número entero."

        # Lógica para formulario de salida
        if tipo_formulario == "salida":
            fecha_salida = data.get("fecha_salida")
            # Generar el nombre de la subcarpeta
            subcarpeta_nombre = generar_nombre_subcarpeta(fecha_salida, nombre_chofer, placa)
            subcarpeta_path = os.path.join(FOTOS_VEHICULOS_DIR, subcarpeta_nombre)

            # Crear la subcarpeta si no existe
            if not os.path.exists(subcarpeta_path):
                os.makedirs(subcarpeta_path)
                logging.info(f"Subcarpeta creada: {subcarpeta_path}")

            # Guardar las fotos de salida en la subcarpeta
            for i in range(1, 5):
                foto_key = f"foto_km_inicial_{i}"
                foto_inicio = files.get(foto_key)
                if foto_inicio:
                    original_extension = os.path.splitext(foto_inicio.filename)[1] if foto_inicio.filename else ".jpg"
                    filename_inicio = f"{subcarpeta_nombre}_salida_{i}{original_extension}"
                    path_inicio = os.path.join(subcarpeta_path, filename_inicio)
                    foto_inicio.save(path_inicio)
                    logging.info(f"Foto de inicio {i} guardada en {path_inicio}")

            # Guardar los datos en el Excel
            fecha_salida_date = datetime.strptime(fecha_salida, "%Y-%m-%d").date()
            fila_salida = [
                fecha_salida_date,  # Fecha
                nombre_chofer,
                data.get("vehiculo"),
                placa,
                fecha_salida_date,  # Fecha de Salida
                data.get("hora_salida"),
                data.get("ubicacion_inicial"),
                data.get("km_inicial"),
                data.get("observaciones_salida"),
                None, None, None, None, None
            ]
            ws_choferes.append(fila_salida)
            wb_choferes.save(REGISTROS_CHOFERES_EXCEL)
            logging.info(f"Datos de salida guardados en nueva fila.")
            return True, "Datos de salida guardados correctamente."

        # Lógica para formulario de llegada
        elif tipo_formulario == "llegada":
            ultimo_registro = None
            for row_idx in range(ws_choferes.max_row, 1, -1):
                if (ws_choferes.cell(row=row_idx, column=2).value == nombre_chofer and
                    ws_choferes.cell(row=row_idx, column=4).value == placa):
                    ultimo_registro = row_idx
                    break

            if ultimo_registro:
                if (ws_choferes.cell(row=ultimo_registro, column=10).value is None and
                    ws_choferes.cell(row=ultimo_registro, column=11).value is None and
                    ws_choferes.cell(row=ultimo_registro, column=12).value is None and
                    ws_choferes.cell(row=ultimo_registro, column=13).value is None and
                    ws_choferes.cell(row=ultimo_registro, column=14).value is None):
                    # Actualizar los datos de llegada en el Excel
                    ws_choferes.cell(row=ultimo_registro, column=10).value = datetime.strptime(data.get("fecha_llegada"), "%Y-%m-%d").date()
                    ws_choferes.cell(row=ultimo_registro, column=11).value = data.get("hora_retorno")
                    ws_choferes.cell(row=ultimo_registro, column=12).value = data.get("ubicacion_final")
                    ws_choferes.cell(row=ultimo_registro, column=13).value = data.get("km_final")
                    ws_choferes.cell(row=ultimo_registro, column=14).value = data.get("observaciones_llegada")
                    wb_choferes.save(REGISTROS_CHOFERES_EXCEL)
                    logging.info(f"Datos de llegada actualizados en fila {ultimo_registro}.")

                    # Obtener la fecha de salida desde el Excel (columna 5: Fecha de Salida)
                    fecha_salida_excel = ws_choferes.cell(row=ultimo_registro, column=5).value
                    if isinstance(fecha_salida_excel, datetime):
                        fecha_salida_str = fecha_salida_excel.strftime("%Y-%m-%d")
                    else:
                        fecha_salida_str = str(fecha_salida_excel)

                    # Generar el nombre de la subcarpeta
                    subcarpeta_nombre = generar_nombre_subcarpeta(fecha_salida_str, nombre_chofer, placa)
                    subcarpeta_path = os.path.join(FOTOS_VEHICULOS_DIR, subcarpeta_nombre)

                    # Crear la subcarpeta si no existe (aunque debería existir desde la salida)
                    if not os.path.exists(subcarpeta_path):
                        os.makedirs(subcarpeta_path)
                        logging.info(f"Subcarpeta creada: {subcarpeta_path}")

                    # Guardar las fotos de llegada en la subcarpeta
                    for i in range(1, 5):
                        foto_key = f"foto_km_final_{i}"
                        foto_fin = files.get(foto_key)
                        if foto_fin:
                            original_extension = os.path.splitext(foto_fin.filename)[1] if foto_fin.filename else ".jpg"
                            filename_fin = f"{subcarpeta_nombre}_llegada_{i}{original_extension}"
                            path_fin = os.path.join(subcarpeta_path, filename_fin)
                            foto_fin.save(path_fin)
                            logging.info(f"Foto de fin {i} guardada en {path_fin} para fila {ultimo_registro}")

                    return True, "Datos de llegada actualizados correctamente.", ultimo_registro
                else:
                    return False, "El último registro ya tiene datos de llegada. No puedes actualizarlo."
            else:
                return False, "No has enviado el Formulario de Datos de Salida correspondiente."

    except Exception as e:
        logging.error(f"Error al procesar datos de choferes: {str(e)}")
        return False, f"Error al procesar datos: {str(e)}"

@app.route('/api/recibir_datos_choferes', methods=['POST'])
def recibir_datos_choferes():
    """Recibe datos o fotos del formulario de choferes."""
    result = procesar_datos_choferes(request.form, request.files)
    if len(result) == 3:  # Caso con row_idx
        success, message, row_idx = result
        if success:
            return jsonify({"status": "success", "message": message, "row_idx": row_idx}), 200
        else:
            return jsonify({"status": "error", "message": message}), 400
    else:  # Caso sin row_idx
        success, message = result
        if success:
            return jsonify({"status": "success", "message": message}), 200
        else:
            return jsonify({"status": "error", "message": message}), 400


@app.route('/api/conductores', methods=['GET'])
def get_conductores():
    """Obtiene la lista de conductores."""
    try:
        if not os.path.exists(CONDUCTORES_CSV_PATH):
            return jsonify({"error": f"No se encontró el archivo de conductores"}), 404

        df = read_csv_with_encoding_detection(CONDUCTORES_CSV_PATH)
        if 'conductor' in df.columns:
            conductores = df['conductor'].dropna().astype(str).tolist()
        else:
            conductores = []
            logging.warning(f"La columna 'conductor' no se encontró en {CONDUCTORES_CSV_PATH}")
        return jsonify(conductores)

    except Exception as e:
        logging.error(f"Error al leer el archivo de conductores: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/vehiculos_info', methods=['GET'])
def get_vehiculos_info():
    """Obtiene la información de los vehículos (tipo y placa)."""
    try:
        if not os.path.exists(VEHICULOS_INFO_CSV_PATH):
            return jsonify({"error": f"No se encontró el archivo de vehículos"}), 404
        df = read_csv_with_encoding_detection(VEHICULOS_INFO_CSV_PATH)
        if 'tipo_vehiculo' in df.columns and 'placa' in df.columns:
            vehiculos = df[['tipo_vehiculo', 'placa']].dropna().astype(str).to_dict('records')
            return jsonify(vehiculos)
        else:
            missing_cols = []
            if 'tipo_vehiculo' not in df.columns:
                missing_cols.append('tipo_vehiculo')
            if 'placa' not in df.columns:
                missing_cols.append('placa')
            logging.warning(f"Faltan columnas en {VEHICULOS_INFO_CSV_PATH}: {', '.join(missing_cols)}")
            return jsonify({"error": f"Faltan columnas: {', '.join(missing_cols)}"}), 400

    except Exception as e:
        logging.error(f"Error al leer el archivo de vehículos: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/descargar-registro-rutas', methods=['GET'])
def descargar_registro_rutas():
    """Descarga el archivo Excel de registros de choferes."""
    try:
        logging.info(f"Intentando enviar archivo de registro de rutas: {REGISTROS_CHOFERES_EXCEL}")
        return send_file(REGISTROS_CHOFERES_EXCEL, as_attachment=True, download_name='registros_choferes.xlsx')
    except Exception as e:
        logging.error(f"Error al generar descarga de Excel de registros de choferes: {str(e)}")
        return str(e), 500

@app.route('/api/listar-carpetas-fotos', methods=['GET'])
def listar_carpetas_fotos():
    """Devuelve la lista de carpetas con el número de fotos en cada una."""
    try:
        carpetas = {}
        for nombre in os.listdir(FOTOS_VEHICULOS_DIR):
            carpeta_path = os.path.join(FOTOS_VEHICULOS_DIR, nombre)
            if os.path.isdir(carpeta_path):
                num_fotos = len([f for f in os.listdir(carpeta_path) if os.path.isfile(os.path.join(carpeta_path, f))])
                carpetas[nombre] = num_fotos
        return jsonify(carpetas)
    except Exception as e:
        logging.error(f"Error al listar carpetas de fotos: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/descargar-carpeta-fotos/<nombre_carpeta>', methods=['GET'])
def descargar_carpeta_fotos_especifica(nombre_carpeta):
    """Comprime y descarga una carpeta específica de fotos_vehiculos."""
    try:
        carpeta_path = os.path.join(FOTOS_VEHICULOS_DIR, nombre_carpeta)
        if not os.path.exists(carpeta_path) or not os.path.isdir(carpeta_path):
            return jsonify({"error": "Carpeta no encontrada"}), 404

        zip_file_path = os.path.join(BASE_DIR, f"{nombre_carpeta}.zip")
        with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, _, files in os.walk(carpeta_path):
                for file in files:
                    zipf.write(os.path.join(root, file),
                              os.path.relpath(os.path.join(root, file),
                                             os.path.join(carpeta_path, '..')))

        # Enviar el archivo ZIP al cliente
        response = send_file(zip_file_path, as_attachment=True, download_name=f"{nombre_carpeta}.zip")

        # Eliminar el archivo ZIP después de enviarlo
        os.remove(zip_file_path)
        logging.info(f"Archivo ZIP temporal eliminado: {zip_file_path}")

        return response
    except Exception as e:
        logging.error(f"Error al comprimir o descargar la carpeta {nombre_carpeta}: {str(e)}")
        return str(e), 500

@app.route('/descargar-carpeta-fotos', methods=['GET'])
def descargar_carpeta_fotos():
    """Comprime y descarga la carpeta de fotos de kilometraje."""
    try:
        zip_file_path = os.path.join(BASE_DIR, "fotos_vehiculos.zip")
        with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, _, files in os.walk(FOTOS_VEHICULOS_DIR):
                for file in files:
                    zipf.write(os.path.join(root, file),
                               os.path.relpath(os.path.join(root, file),
                                               os.path.join(FOTOS_VEHICULOS_DIR, '..')))
        logging.info(f"Intentando enviar carpeta de fotos comprimida: {zip_file_path}")
        return send_file(zip_file_path, as_attachment=True, download_name='fotos_vehiculos.zip')
    except Exception as e:
        logging.error(f"Error al comprimir o descargar la carpeta de fotos: {str(e)}")
        return str(e), 500


# API endpoints para el sistema de logística
@app.route('/api/logistica/materiales', methods=['GET'])
def obtener_materiales_logistica():
    """Devuelve la lista de materiales desde el archivo CSV de logística."""
    try:
        if os.path.exists(LOGISTICA_MATERIALES_CSV_PATH):
            df = read_csv_with_encoding_detection(LOGISTICA_MATERIALES_CSV_PATH)
            materiales = df[['material', 'unidad']].to_dict('records')
            return jsonify(materiales)
        else:
            return jsonify({"error": "Archivo de materiales de logística no encontrado"}), 404
    except Exception as e:
        logging.exception(f"Error al obtener materiales de logística: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/logistica/clientes', methods=['GET'])
def obtener_clientes_logistica():
    """Devuelve la lista de clientes desde el archivo CSV de logística."""
    try:
        if os.path.exists(LOGISTICA_CLIENTES_CSV_PATH):
            df = read_csv_with_encoding_detection(LOGISTICA_CLIENTES_CSV_PATH)
            clientes = df['cliente'].tolist()
            return jsonify(clientes)
        else:
            return jsonify({"error": "Archivo de clientes de logística no encontrado"}), 404
    except Exception as e:
        logging.exception(f"Error al obtener clientes de logística: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/logistica/enviar-requerimientos', methods=['POST'])
def recibir_requerimientos_logistica():
    """Recibe los datos de requerimientos desde la app Android de logística."""
    try:
        datos = request.json
        logging.info(f"Datos de logística recibidos: {datos}")

        if procesar_logistica_requerimientos(datos):
            return jsonify({"status": "success", "message": "Requerimientos procesados correctamente"}), 200
        else:
            return jsonify({"status": "error", "message": "Error al procesar requerimientos"}), 500
    except Exception as e:
        logging.exception(f"Error al recibir requerimientos de logística: {str(e)}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/logistica/descargar-requerimientos', methods=['GET'])
def descargar_requerimientos_logistica():
    """Descarga el archivo Excel de requerimientos de logística."""
    return descargar_logistica_excel_flask()

@app.route('/api/logistica/descargar-bdd', methods=['GET'])
def descargar_bdd_logistica():
    """Descarga el archivo CSV de la base de datos de materiales de logística."""
    return descargar_bdd_logistica_flask()

@app.route('/api/logistica/subir-bdd', methods=['POST'])
def subir_bdd_logistica():
    """Sube (actualiza) el archivo CSV de la base de datos de materiales de logística."""
    if 'file' not in request.files:
        logging.warning("No se encontró 'file' en la solicitud de subida de BDD.")
        return jsonify({"error": "No se encontró el archivo en la solicitud"}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        logging.warning("Nombre de archivo vacío en la solicitud de subida de BDD.")
        return jsonify({"error": "Nombre de archivo vacío"}), 400
    
    if file and file.filename.endswith('.csv'):
        try:
            # Guardar el archivo, sobrescribiendo el existente
            file.save(LOGISTICA_MATERIALES_CSV_PATH)
            logging.info(f"Archivo BDD de logística '{file.filename}' subido y guardado como '{LOGISTICA_MATERIALES_CSV_PATH}'")
            return jsonify({"status": "success", "message": "Base de datos de materiales actualizada correctamente."}), 200
        except Exception as e:
            logging.error(f"Error al guardar el archivo BDD de logística subido: {str(e)}")
            return jsonify({"error": f"Error al guardar el archivo en el servidor: {str(e)}"}), 500
    else:
        logging.warning(f"Archivo no válido o tipo incorrecto para subida de BDD: {file.filename}")
        return jsonify({"error": "Archivo no válido o tipo incorrecto. Se esperaba un archivo .csv"}), 400

@app.route('/api/logistica/descargar-clientes', methods=['GET'])
def descargar_clientes_logistica():
    """Permite descargar el archivo CSV de clientes."""
    try:
        if os.path.exists(LOGISTICA_CLIENTES_CSV_PATH):
            return send_file(LOGISTICA_CLIENTES_CSV_PATH, as_attachment=True, download_name="logistica_clientes.csv")
        else:
            return jsonify({"error": "Archivo de clientes no encontrado"}), 404
    except Exception as e:
        logging.exception(f"Error al descargar archivo de clientes: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/logistica/subir-clientes', methods=['POST'])
def subir_clientes_logistica():
    """Sube (actualiza) el archivo CSV de clientes de logística."""
    if 'file' not in request.files:
        logging.warning("No se encontró 'file' en la solicitud de subida de clientes.")
        return jsonify({"error": "No se encontró el archivo en la solicitud"}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        logging.warning("Nombre de archivo vacío en la solicitud de subida de clientes.")
        return jsonify({"error": "Nombre de archivo vacío"}), 400
    
    if file and file.filename.endswith('.csv'):
        try:
            # Guardar el archivo, sobrescribiendo el existente
            file.save(LOGISTICA_CLIENTES_CSV_PATH)
            logging.info(f"Archivo de clientes '{file.filename}' subido y guardado como '{LOGISTICA_CLIENTES_CSV_PATH}'")
            return jsonify({"status": "success", "message": "Base de datos de clientes actualizada correctamente."}), 200
        except Exception as e:
            logging.error(f"Error al guardar el archivo de clientes subido: {str(e)}")
            return jsonify({"error": f"Error al guardar el archivo en el servidor: {str(e)}"}), 500
    else:
        logging.warning(f"Archivo no válido o tipo incorrecto para subida de clientes: {file.filename}")
        return jsonify({"error": "Archivo no válido o tipo incorrecto. Se esperaba un archivo .csv"}), 400

# API endpoints para AEM (Área de Equipos y Maquinarias)
@app.route('/api/aem/descargar-bdd/<tipo_bdd>', methods=['GET'])
def descargar_bdd_aem(tipo_bdd):
    """Descarga los archivos CSV de base de datos para AEM (conductores y vehículos)."""
    try:
        archivos_bdd_aem = {
            'conductores': CONDUCTORES_CSV_PATH,
            'vehiculos': VEHICULOS_INFO_CSV_PATH
        }

        if tipo_bdd not in archivos_bdd_aem:
            return jsonify({"error": "Tipo de base de datos no válido. Use 'conductores' o 'vehiculos'"}), 400

        archivo_path = archivos_bdd_aem[tipo_bdd]

        if not os.path.exists(archivo_path):
            logging.error(f"Archivo BDD AEM no encontrado: {archivo_path}")
            return jsonify({"error": f"Archivo de base de datos {tipo_bdd} no encontrado en el servidor."}), 404

        # Nombres de descarga personalizados
        nombres_descarga = {
            'conductores': 'aem_conductores.csv',
            'vehiculos': 'aem_vehiculos.csv'
        }

        logging.info(f"Intentando enviar archivo BDD AEM ({tipo_bdd}): {archivo_path}")
        return send_file(archivo_path, as_attachment=True, download_name=nombres_descarga[tipo_bdd])

    except Exception as e:
        logging.error(f"Error al generar descarga de CSV BDD AEM ({tipo_bdd}): {str(e)}")
        return jsonify({"error": f"Error interno del servidor: {str(e)}"}), 500

@app.route('/api/aem/subir-bdd/<tipo_bdd>', methods=['POST'])
def subir_bdd_aem(tipo_bdd):
    """Sube (actualiza) los archivos CSV de base de datos para AEM."""
    try:
        archivos_bdd_aem = {
            'conductores': CONDUCTORES_CSV_PATH,
            'vehiculos': VEHICULOS_INFO_CSV_PATH
        }

        if tipo_bdd not in archivos_bdd_aem:
            return jsonify({"error": "Tipo de base de datos no válido. Use 'conductores' o 'vehiculos'"}), 400

        if 'file' not in request.files:
            logging.warning(f"No se encontró 'file' en la solicitud de subida de BDD AEM ({tipo_bdd}).")
            return jsonify({"error": "No se encontró el archivo en la solicitud"}), 400

        file = request.files['file']

        if file.filename == '':
            logging.warning(f"Nombre de archivo vacío en la solicitud de subida de BDD AEM ({tipo_bdd}).")
            return jsonify({"error": "Nombre de archivo vacío"}), 400

        if file and file.filename.endswith('.csv'):
            archivo_path = archivos_bdd_aem[tipo_bdd]
            file.save(archivo_path)
            logging.info(f"Archivo BDD AEM ({tipo_bdd}) '{file.filename}' subido y guardado como '{archivo_path}'")
            return jsonify({"status": "success", "message": f"Base de datos de {tipo_bdd} AEM actualizada correctamente."}), 200
        else:
            logging.warning(f"Archivo no válido o tipo incorrecto para subida de BDD AEM ({tipo_bdd}): {file.filename}")
            return jsonify({"error": "Archivo no válido o tipo incorrecto. Se esperaba un archivo .csv"}), 400

    except Exception as e:
        logging.error(f"Error al guardar el archivo BDD AEM ({tipo_bdd}) subido: {str(e)}")
        return jsonify({"error": f"Error al guardar el archivo en el servidor: {str(e)}"}), 500

# API endpoints para gestión de bases de datos CSV de operaciones
@app.route('/api/operaciones/descargar-bdd/<tipo_bdd>', methods=['GET'])
def descargar_bdd_operaciones(tipo_bdd):
    """Descarga los archivos CSV de base de datos de operaciones."""
    try:
        archivos_bdd = {
            'materiales': MATERIALES_CSV_PATH,
            'equipos': EQUIPOS_CSV_PATH,
            'vehiculos': VEHICULOS_CSV_PATH,
            'personal': PERSONAL_CSV_PATH
        }

        if tipo_bdd not in archivos_bdd:
            return jsonify({"error": "Tipo de base de datos no válido"}), 400

        archivo_path = archivos_bdd[tipo_bdd]

        if not os.path.exists(archivo_path):
            logging.error(f"Archivo BDD operaciones no encontrado: {archivo_path}")
            return jsonify({"error": f"Archivo de base de datos {tipo_bdd} no encontrado en el servidor."}), 404

        # Nombres de descarga personalizados
        nombres_descarga = {
            'materiales': 'operaciones_materiales.csv',
            'equipos': 'operaciones_equipos.csv',
            'vehiculos': 'operaciones_vehiculos.csv',
            'personal': 'operaciones_personal.csv'
        }

        logging.info(f"Intentando enviar archivo BDD de operaciones ({tipo_bdd}): {archivo_path}")
        return send_file(archivo_path, as_attachment=True, download_name=nombres_descarga[tipo_bdd])

    except Exception as e:
        logging.error(f"Error al generar descarga de CSV BDD de operaciones ({tipo_bdd}): {str(e)}")
        return jsonify({"error": f"Error interno del servidor: {str(e)}"}), 500

@app.route('/api/operaciones/subir-bdd/<tipo_bdd>', methods=['POST'])
def subir_bdd_operaciones(tipo_bdd):
    """Sube (actualiza) los archivos CSV de base de datos de operaciones."""
    try:
        archivos_bdd = {
            'materiales': MATERIALES_CSV_PATH,
            'equipos': EQUIPOS_CSV_PATH,
            'vehiculos': VEHICULOS_CSV_PATH,
            'personal': PERSONAL_CSV_PATH
        }

        if tipo_bdd not in archivos_bdd:
            return jsonify({"error": "Tipo de base de datos no válido"}), 400

        if 'file' not in request.files:
            logging.warning(f"No se encontró 'file' en la solicitud de subida de BDD operaciones ({tipo_bdd}).")
            return jsonify({"error": "No se encontró el archivo en la solicitud"}), 400

        file = request.files['file']

        if file.filename == '':
            logging.warning(f"Nombre de archivo vacío en la solicitud de subida de BDD operaciones ({tipo_bdd}).")
            return jsonify({"error": "Nombre de archivo vacío"}), 400

        if file and file.filename.endswith('.csv'):
            archivo_path = archivos_bdd[tipo_bdd]
            file.save(archivo_path)
            logging.info(f"Archivo BDD de operaciones ({tipo_bdd}) '{file.filename}' subido y guardado como '{archivo_path}'")
            return jsonify({"status": "success", "message": f"Base de datos de {tipo_bdd} actualizada correctamente."}), 200
        else:
            logging.warning(f"Archivo no válido o tipo incorrecto para subida de BDD operaciones ({tipo_bdd}): {file.filename}")
            return jsonify({"error": "Archivo no válido o tipo incorrecto. Se esperaba un archivo .csv"}), 400

    except Exception as e:
        logging.error(f"Error al guardar el archivo BDD de operaciones ({tipo_bdd}) subido: {str(e)}")
        return jsonify({"error": f"Error al guardar el archivo en el servidor: {str(e)}"}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False) # debug=True para desarrollo