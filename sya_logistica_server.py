# sya_logistica_server.py
import os
import logging
from datetime import datetime
import pandas as pd
import openpyxl
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("server_log.log"),
        logging.StreamHandler()
    ]
)

# Use absolute paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "sya_logistica_requerimientos.xlsx")
MATERIALES_CSV_PATH = os.path.join(BASE_DIR, "logistica_materiales.csv")

def crear_excel_si_no_existe():
    """Crea el archivo Excel si no existe."""
    if not os.path.exists(EXCEL_FILE):
        logging.info(f"Creando archivo Excel: {EXCEL_FILE}")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Requerimientos"
        
        # Definir cabeceras
        cabeceras = [
            "Fecha", "Solicitante", "Orden de Trabajo", "Cliente", 
            "Producto", "Unidad", "Cantidad", "Stock", "Adquirido", 
            "Saldo", "Observaciones"
        ]
        
        for col_num, header in enumerate(cabeceras, 1):
            ws.cell(row=1, column=col_num).value = header
        
        wb.save(EXCEL_FILE)
        logging.info("Archivo Excel creado exitosamente")

def procesar_requerimientos(datos):
    """Procesa los datos de requerimientos y los guarda en el Excel."""
    try:
        # Asegurar que el archivo Excel existe
        crear_excel_si_no_existe()
        
        # Cargar el archivo Excel existente
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["Requerimientos"]
        
        # Obtener la última fila con datos
        ultima_fila = ws.max_row
        
        # Datos comunes para todos los productos
        fecha = datos.get('fecha', '')
        solicitante = datos.get('solicitante', '')
        orden_trabajo = datos.get('orden_trabajo', '')
        cliente = datos.get('cliente', '')
        
        # Procesar cada producto en la lista
        productos = datos.get('productos', [])
        for producto in productos:
            # Incrementar el número de fila
            ultima_fila += 1
            
            # Insertar datos en la fila
            ws.cell(row=ultima_fila, column=1).value = fecha
            ws.cell(row=ultima_fila, column=2).value = solicitante
            ws.cell(row=ultima_fila, column=3).value = orden_trabajo
            ws.cell(row=ultima_fila, column=4).value = cliente
            ws.cell(row=ultima_fila, column=5).value = producto.get('producto', '')
            ws.cell(row=ultima_fila, column=6).value = producto.get('unidad', '')
            ws.cell(row=ultima_fila, column=7).value = producto.get('cantidad', 0.0)
            # Las columnas 8, 9, 10 y 11 (Stock, Adquirido, Saldo y Observaciones) se dejan vacías
        
        # Guardar el archivo Excel
        wb.save(EXCEL_FILE)
        logging.info(f"Requerimientos recibidos de {solicitante} procesados exitosamente")
        return True
    except Exception as e:
        logging.exception(f"Error al procesar requerimientos: {str(e)}")
        return False

@app.route('/api/materiales', methods=['GET'])
def obtener_materiales():
    """Devuelve la lista de materiales desde el archivo CSV."""
    try:
        if os.path.exists(MATERIALES_CSV_PATH):
            df = pd.read_csv(MATERIALES_CSV_PATH)
            materiales = df[['material', 'unidad']].to_dict('records')
            return jsonify(materiales)
        else:
            return jsonify({"error": "Archivo de materiales no encontrado"}), 404
    except Exception as e:
        logging.exception(f"Error al obtener materiales: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/enviar-requerimientos', methods=['POST'])
def recibir_requerimientos():
    """Recibe los datos de requerimientos desde la app Android."""
    try:
        datos = request.json
        logging.info(f"Datos recibidos: {datos}")
        
        if procesar_requerimientos(datos):
            return jsonify({"status": "success", "message": "Requerimientos procesados correctamente"}), 200
        else:
            return jsonify({"status": "error", "message": "Error al procesar requerimientos"}), 500
    except Exception as e:
        logging.exception(f"Error al recibir requerimientos: {str(e)}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/descargar-requerimientos', methods=['GET'])
def descargar_requerimientos():
    """Descarga el archivo Excel de requerimientos."""
    try:
        if not os.path.exists(EXCEL_FILE):
            crear_excel_si_no_existe()
            
        return send_file(EXCEL_FILE, as_attachment=True, download_name="sya_logistica_requerimientos.xlsx")
    except Exception as e:
        logging.exception(f"Error al descargar requerimientos: {str(e)}")
        return jsonify({"status": "error", "message": str(e)}), 500

if __name__ == '__main__':
    # Asegurar que el archivo Excel existe al iniciar el servidor
    crear_excel_si_no_existe()
    app.run(host='0.0.0.0', port=5000, debug=True)
